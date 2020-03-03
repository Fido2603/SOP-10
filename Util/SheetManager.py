import openpyxl
import os

# Load workbook
if os.path.exists('data.xlsx'):
    wb = openpyxl.load_workbook('data.xlsx')
else:
    wb = openpyxl.Workbook()
    wb.save('data.xlsx')


def get_worksheet(name=None):
    if name is None or not isinstance(name, str):
        ws = wb.active
    else:
        ws = wb.get_sheet_by_name(name)
    return ws


def new_worksheet(name=None):
    # Create new worksheet
    if name is None or not isinstance(name, str):
        ws = wb.create_sheet()
    else:
        ws = wb.create_sheet(name)

    # Fill out first columns
    ws["A1"] = "Input"
    ws["B1"] = "Avg. time"
    ws["C1"] = "Min. time"
    ws["D1"] = "Max. time"
    ws["E1"] = "Output"

    # Return the newly created worksheet
    return ws


def write_newentry(entry, ws=None):
    if ws is None:
        ws = wb.active

    # Get a cell that isn't used yet
    row = 1
    while True:
        cell = ws.cell(column=1, row=row)
        if cell.value is None:
            break
        row += 1

    # Write entry to cells
    column = 1
    for value in entry:
        if column == 1:
            cell.value = value
        else:
            ws.cell(column=column, row=row, value=value)
        column += 1

    # Save
    save()


def save():
    wb.save('data.xlsx')
